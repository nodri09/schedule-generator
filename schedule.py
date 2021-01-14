"""
Schedule Generating file
Created by Gelovani Nodar
"""
import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string
from set_calc import yeardays, get_shiftset, Manager, printff
from sys import argv, exit

# Check if command line arguments are 3
# Check if second and third files have 'xlsx' file extention
# If not correct any of the above, exit
"""
if len(argv) == 3:
    if argv[1].endswith('xlsx') and argv[2].endswith('xlsx'):
      printff("Starting the program...\n")
    else:
      printff("Incorrect file extention\n")
      exit()
else:
    printff("Something went wrong. Cannot open the file.\n")
    exit()
"""

def main(manager_file, yearInput, monthInput, teams_file):


    # Transfer data from Excel file
    # Open Excel Workbook
    printff(f"Opening file: {manager_file}\n")
    wb = xl.load_workbook(manager_file)

    # Open Active worksheet
    printff("Opening active worksheet...\n")
    sheet = wb.active

    # List of Manager class - All the managers in one List
    managers = []

    # Store amount of rows in one variable
    max_row = sheet.max_row

    # Transfer information from Managers excel
    printff("Retreiving information from columns...\n")
    for row in range(2, max_row + 1):
        name = sheet['A' + str(row)].value
        shift_type = sheet['B' + str(row)].value
        mset = sheet['C' + str(row)].value
        email = sheet['D' + str(row)].value

        # Put Retreived info into the Manager class
        manager = Manager(name, shift_type, mset, email)

        # Append every manager into the managers List
        managers.append(manager)

    # Ask user to provide Year and Month
    # Check to be numbers only
    printff("Dictate the year:\n ")
    user_year = yearInput

    while False:
      if user_year == None or type(user_year) != int():
        False
      else:
        True

    printff("Dictate the month:\n ")
    user_month = monthInput

    while False:
      if user_month == None or type(user_month) != int():
        False
      else:
        True

    year = user_year
    printff(f"Year is: {year}\n")
    month = user_month
    printff(f"Month is: {month}\n")

    # print("\nJan.'21, Set 1")
    # print([str(day) for day in yeardays(year) if day.month == month and get_shiftset(day)=="Set 1"])
    # print("Jan.'21, Set 2")
    # print([str(day) for day in yeardays(year) if day.month == month and get_shiftset(day)=="Set 2"])

    # Generate set days of the given month of the given year
    # Assign generated set-days to managers, which have the same set
    printff("Generating set days for each Manager\n")
    for manager in managers:
        for day in yeardays(year):
            if day.month == month and get_shiftset(day) == "Set 1" and manager.mset == 1:
                manager.shifts.append(day)
            elif day.month == month and get_shiftset(day) == "Set 2" and manager.mset == 2:
                manager.shifts.append(day)

    # Change Python Date to Excel Date
    for manager in managers:
        for i in range(len(manager.shifts)):
            manager.shifts[i] = str(manager.shifts[i]).replace("-", "/")


    # Open Team's schedule template file
    wb1 = xl.load_workbook(teams_file)
    printff(f"Opening file: {teams_file}\n")

    # Load 'Shifts' sheet
    tsheet = wb1['Shifts']

    # Translateing managers info to excel column
    printff("Translateing managers info to excel column\n")
    for manager in managers:
        amount = len(manager.shifts)
        tsheet.insert_rows(idx=2, amount=amount + 1)
        for row in range(2, amount + 2):
            tsheet['A' + str(row)] = manager.name
            tsheet['B' + str(row)] = manager.email
            tsheet['D' + str(row)] = manager.shifts[row - 2]
            tsheet['F' + str(row)] = manager.shifts[row - 2]

            if manager.shift_type.lower() == 'morning':
                tsheet['E' + str(row)] = '07:00'
                tsheet['G' + str(row)] = '15:00'
                tsheet['H' + str(row)] = '2. Blue'
            elif manager.shift_type.lower() == 'afternoon':
                tsheet['E' + str(row)] = '15:00'
                tsheet['G' + str(row)] = '23:00'
                tsheet['H' + str(row)] = '3. Green'
            elif manager.shift_type.lower() == 'night':
                tsheet['E' + str(row)] = '00:00'
                tsheet['G' + str(row)] = '07:00'
                tsheet['H' + str(row)] = '5. Pink'

    # Save and close Workbooks
    printff(f"Saving {teams_file}\n")
    wb1.save(teams_file)

    printff(f"Closing {manager_file}\n")
    wb.close()

    printff(f"Closing {teams_file}\n")
    wb1.close()
