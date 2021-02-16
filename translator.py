"""
Translator App from Excel to Microsoft Teams Shifts template file
"""

import openpyxl as xl
from set_calc import Manager
import datetime as dt


wb = xl.load_workbook("schedule.xlsx", data_only=True) # Open Workbook

try:
    sheet = wb["schedule"] # Open active sheet
except:
    print("was not able to open sheet")

max_row = sheet.max_row # Maximum rows of the sheet

managers = {} # Dictionary of Managers class

# Iterate over the rows in the specific column
# Create manager's class for every manager in excel file
# Put all Managers into managers list
for row in range(3, sheet.max_row+1):
    for column in "A":
        cell_name = "{}{}".format(column, row)
        cell_color = sheet[cell_name].fill.start_color.index
        manager = Manager(sheet[cell_name].value)
        manager.color = cell_color
        managers.update({manager:row})

for col in sheet.iter_cols(min_col=2, max_col=2, min_row=3, max_row=sheet.max_row):
    for cell in col:
        for manager, row in managers.items():
            if cell.row == row:
                manager.email = cell.value


cells = [] # list that stores all the cells except date and managers name
excel_dates = {} # Dictionary of the dates that should be translated

# Update excel_dates with data from excel
# Key is the Column letter of the excel column
# Value is the value of the excel file's cell
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=sheet.max_column):
    for cell in row:
        if cell.row == 2:
            excel_dates.update({cell.column_letter:str(cell.value)})
        elif cell.value != None or cell.value != 0:
            cells.append(cell)


copied_cells = []
# For every cell with none value assing value to 0
# for better manipulation on the data
for cell in cells:
    if cell.value != 0:
        copied_cells.append(cell)


# Update manager's shifts list with role and column letter.
# Later letter will be changed to actual date
# Key is the value, a.k.a FS role
# Value is the Column letter of the excel column
for cell in copied_cells:
    for manager, indx in managers.items():
        if cell.row == indx:
            manager.shifts.append(dict({cell.value:cell.column_letter}))

# Get current year
current_year = dt.datetime.now().year


# Translate excel dates into shifts tamplate dates
for key, value in excel_dates.items():
    value = value.replace(".", "/")
    value = value + "/" + str(current_year)

    y = dt.datetime.strptime(value, "%d/%m/%Y").date()
    value = dt.datetime.strftime(y, "%-m/%d/%Y")

    excel_dates.update({key:value})



for manager in managers:
    for shift in manager.shifts:
        for role, cl_lett in shift.items():
            for column_letter, dates in excel_dates.items():
                if cl_lett == column_letter:
                    cl_lett = dates
                    shift.update({role:cl_lett})

"""
for manager in managers:
    for shift in manager.shifts:
        for role, cdate in shift.items():
            print(manager.name, manager.color, role, cdate)




# Delete used/unnecessary information
cells.clear()
excel_dates.clear()

wb1 = xl.load_workbook('team_file.xlsx')

ws = wb1['Shifts'] # Load 'Shifts' sheet

# y = dt.datetime.strptime(x, "%d/%m/%Y").date() + dt.timedelta(days=1)

for manager in managers:
    for shift in manager.shifts:
        ws.insert_rows(idx=2, amount=1)
        ws['A' + str(2)] = manager.name
        ws['B' + str(2)] = manager.email

        if manager.color == 8:
            ws['E' + str(2)] = "07:00"
            ws['G' + str(2)] = "19:00"
            ws['H' + str(2)] = "8. DarkBlue"

            for role, shift_date in shift.items():
                ws['D' + str(2)] = shift_date
                ws['F' + str(2)] = shift_date
                ws['I' + str(2)] = role

        elif manager.color == 7:
            ws['E' + str(2)] = "19:00"
            ws['G' + str(2)] = "07:00"
            ws['H' + str(2)] = "12. DarkYellow"

            for role, shift_date in shift.items():
                y = dt.datetime.strptime(shift_date, "%m/%d/%Y").date() + dt.timedelta(days=1)

                ws['D' + str(2)] = shift_date
                ws['F' + str(2)] = dt.datetime.strftime(y, "%-m/%d/%Y")
                ws['I' + str(2)] = role


wb1.save('team_file.xlsx')
wb1.close()
"""

wb.close()